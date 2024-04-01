# encoding:utf-8
import os
import openpyxl
import re
import shutil
import pandas
from common.random_tool import RandomData
from common import ROOT_PATH

random_name = 'random'  # 随机函数开始名称规则
random_func_list = [func for func in dir(RandomData) if random_name in func]  # 以random命名的随机函数集合列表


class ExcelTool:
    """读取excel文件中的数据"""

    def __init__(self, file, sheet="", create=False):
        """
        :param file:excel文件路径
        :param sheet:sheet工作页名
        :param create:创建excel的初始化,当为True时,文件路径下创建文件
        """
        self.file = file
        self.create = create
        if self.create:
            self.wb = openpyxl.Workbook()
            self.sheet = self.wb.active
        else:
            self.wb = openpyxl.load_workbook(self.file)
            if sheet:
                self.sheet = self.wb[sheet]

    def read_excel(self):
        """读取excel文件"""
        res = [
            dict(
                zip(
                    list(map(lambda x: x.value, next(self.sheet.rows))),
                    list(map(lambda x: x.value, m)),
                )
            )
            for m in list(self.sheet.rows)[1:]
        ]
        return res

    def readExcel(self, sheet_name=""):
        """读取excel文件，并返回字段（若excel存在空值，则忽略）"""
        result = []
        if sheet_name:
            table = self.wb[sheet_name]
        else:
            table = self.wb.sheetnames[0]
        rows, cols = table.max_row, table.max_column
        all_title = [table.cell(1, i + 1).value for i in range(cols)]
        for row in range(1, rows):
            titles = all_title[:]
            values = []
            for col in range(cols):
                data = table.cell(row + 1, col + 1).value
                # 值为空，则忽略并删除字典的key
                if data:
                    values.append(data)
                else:
                    del titles[len(values)]
            result.append(dict(zip(titles, values)))
        return result

    def write_excel(self, row, column, value):
        """
        写入excel
        :param row:插入单元格所在行
        :param column:插入单元格所在列
        :param value:插入单元格的值
        """
        self.sheet.cell(row, column).value = value
        self.wb.save(self.file)

    def write_row_data(self, row, *args, **kwargs):
        """
        根据传入元素填充数据行
        :param row:传入数据所在行
        :param args:传入数据
        """
        value_dict = {"begin_row": 1, "title": []}
        value_dict.update(kwargs)
        if self.create:
            for index, value in enumerate(value_dict["title"]):
                self.sheet.cell(1, index + value_dict["begin_row"]).value = value
        for index, value in enumerate(args):
            self.sheet.cell(row + 1, index + value_dict["begin_row"]).value = value

    def excel_save(self):
        """保存excel"""
        self.wb.save(self.file)

    def read_random_excel(self, sheet_name):
        """读取excel，默认读取第一个sheet"""
        result = []
        table = self.wb[sheet_name]
        rows, cols = table.max_row, table.max_column
        all_title = [table.cell(1, i + 1).value for i in range(cols)]
        for row in range(1, rows):
            titles = all_title[:]
            values = []
            for col in range(cols):
                data = str(table.cell(row + 1, col + 1).value)
                # 值为空，则忽略并删除字典的key
                if not data:
                    del titles[len(values)]
                # 如果有随机函数
                elif "random" in data:
                    # 查找匹配的数据生成函数并替换值
                    for func in random_func_list:
                        re_match = func + "\\(.*?\\)"
                        re_list = re.findall(re_match, data)
                        if re_list:
                            data = data.replace(
                                re_list[0], str(eval("RandomData()." + re_list[0]))
                            )
                            values.append(data)
                            break
                # 其它值
                else:
                    values.append(data)
            result.append(dict(zip(titles, values)))
        return result


class ExcelToolsPandas:
    """
    作者：李国彬
    用于读取xls，xlsx数据，用于生成随机数据，文件合并，文件复制等操作
    """

    def __init__(self, sys_name: str = "djw"):
        """默认为大教务系统"""
        self.sys_name = sys_name

    @staticmethod
    def _read_excel_data_frame(file: str, sheet_name="", ignore_num: int = 0):
        """读取excel文件对应sheet页签返回DataFrame结构数据"""
        if not os.path.isfile(file):
            raise FileNotFoundError(f'{file}文件未找到')
        ignore_num_list = [i for i in range(ignore_num)] if ignore_num else None
        if sheet_name:  # 有sheet_name则读取对应对应sheet
            excel_header = pandas.read_excel(io=file, sheet_name=sheet_name, skiprows=ignore_num_list)
        else:  # 没有则读取第一个
            excel_header = pandas.read_excel(io=file, skiprows=ignore_num_list)
        header_name = list(excel_header)
        convert = {i: str for i in header_name}  # 将所有列均按str读取处理
        if sheet_name:  # 有sheet_name则读取对应对应sheet
            excel = pandas.read_excel(
                io=file,
                sheet_name=sheet_name,
                skiprows=ignore_num_list,
                converters=convert,
                keep_default_na=False,
            )
        else:  # 没有则读取第一个
            excel = pandas.read_excel(
                io=file,
                skiprows=ignore_num_list,
                keep_default_na=False,
                converters=convert,
            )
        return excel

    def _read_excel_fun_data(self, file: str, sheet_name="", ignore_num: int = 0):
        """读取excel文件将随机函数执行，重新生成数据，返回DataFrame结构数据"""
        re_match = f"{random_name}.*?\\(.*?\\)"  # 随机函数正则匹配
        data_frame = self._read_excel_data_frame(file=file, sheet_name=sheet_name, ignore_num=ignore_num)
        header_names = list(data_frame)
        for header in header_names:
            values = data_frame[header].values
            for i in range(len(values)):
                # 判断是否可能有随机函数
                values[i] = str(values[i])
                if random_name in values[i]:
                    re_list = set(re.findall(re_match, values[i]))
                    if re_list:  # 判断匹配结果不为空
                        for j in re_list:
                            try:
                                values[i] = values[i].replace(j, str(eval("RandomData()." + j)))
                            except Exception as e:
                                raise e
            # 替换为随机函数后的数据
            data_frame[header] = values
        return data_frame

    def read_excel(self, file: str, sheet_name="", ignore_num: int = None):
        """只读取excel中的数据，并自动去除为空的数据,返回列表数据"""
        # 获取excel列头
        dataframe = self._read_excel_data_frame(
            file=file,
            sheet_name=sheet_name,
            ignore_num=ignore_num
        )
        data_list = dataframe.to_dict("records")
        # 剔除空数据
        for i in data_list:
            keys = list(i.keys())
            for key in keys:
                if not i[key]:
                    del i[key]
        return data_list

    def read_excel_func(self, file, sheet_name="", ignore_num=0):
        """读取excel数据并将随机函数转化为数据，并去掉为空的数据，返回列表数据"""
        data_frame = self._read_excel_fun_data(file=file, sheet_name=sheet_name, ignore_num=ignore_num)
        data_frame.fillna('', inplace=True)
        data = data_frame.to_dict("records")
        # 剔除空数据
        for i in data:
            keys = list(i.keys())
            for key in keys:
                if i[key] == '':
                    del i[key]
        return data

    def read_excel_func_save(self, file, sheet_name="", ignore_num: int = 0):
        """读取xls,xlsx数据，并将数据中的随机函数执行，并返回生成一个新的文件"""
        path = os.path.dirname(file)  # file同级文件目录
        file_name = "随机_" + os.path.basename(file)
        save_file = os.path.join(path, file_name)
        shutil.copyfile(file, save_file)  # 复制文件
        data = self._read_excel_fun_data(file=file, sheet_name=sheet_name, ignore_num=ignore_num)
        try:
            # excel非xls文件读取
            pd_writer = pandas.ExcelWriter(save_file, mode='a', engine='openpyxl', if_sheet_exists='overlay')
        except OSError:
            # excel为xls文件读取
            pd_writer = pandas.ExcelWriter(save_file)
        data.to_excel(excel_writer=pd_writer, index=False, startrow=ignore_num)
        pd_writer.save()
        pd_writer.close()
        return save_file

    def read_excels_combine_save(
            self, file1: str, conditions: list, file2: str = "", ignore_num_file1=0
    ):
        """
        读取excel文件，并根据conditions规则处理file1文件数据，并生成新的文件
        作者：李国彬
        :param file1:需要填充数据的主文件
        :param conditions: 填充数据的规则
        例如：[{'文件1字段1': 'test'}, {'文件1字段1': ['文件2字段1', '文件1字段2', '文件2字段2']}, {'文件1字段1': {'文件2字段2': [1, 2]}}]
        {'文件1字段1': 'test'} 表示：文件1的字段1填充值均为test
        {'文件1字段1': ['文件2字段1', '文件1字段2', '文件2字段1']} 表示：从文件1字段2和文件2字段2进行匹配，值相同则把文件2字段1填充到文件1字段1
        {'文件1字段1': {'文件2字段2': [1, 2]}} 表示：文件1字段1填充3个值（列表总和）分别取文件2字段2第一个值1次，第二个值2次（列表索引的顺序）
        :param file2:获取数据第二个文件
        :param ignore_num_file1:第一个文件中忽略的首行
        """
        # 读取文件1
        file1_data_frame = self._read_excel_data_frame(file=file1, ignore_num=ignore_num_file1)
        file1_headers = list(file1_data_frame)  # 列头字段
        # 读取文件2
        if file2:
            file2_data_frame = self._read_excel_data_frame(file2)
        else:
            file2_data_frame = {}
        for i in conditions:
            if isinstance(i, dict):  # 判断conditions数据结构是否正确
                item = list(i.items())
                key = item[0][0]
                value = item[0][1]
                # 直接填充某列值
                if isinstance(value, str):
                    file1_data_frame[key] = value
                elif isinstance(value, int):
                    file1_data_frame[key].fillna(0, inplace=True)
                    file1_data_frame[key] = value
                elif isinstance(value, float):
                    file1_data_frame[key].fillna(0.1, inplace=True)
                    file1_data_frame[key] = value
                elif file2_data_frame.items():  # 判断文件2是否有上传，没有则抛出异常
                    # 根据匹配主键自动填充对应列值
                    if isinstance(value, list) and len(value) == 3:
                        match_key, file1_key, file2_key = value[0], value[1], value[2]
                        # 根据匹配条件合并两个文件
                        file1_data_frame = file1_data_frame.merge(
                            file2_data_frame,
                            how="left",
                            left_on=file1_key,
                            right_on=file2_key,
                        )
                        # 删除合并后的空列（该列为file1填充值的列）
                        file1_data_frame.drop(columns=[key], inplace=True, axis=1)
                        # 将合并后匹配值行列表修改为填充值的列
                        file1_data_frame.rename(columns={match_key: key}, inplace=True)
                        file1_data_frame = file1_data_frame[file1_headers]
                    # 按顺序填充对应数量的列值
                    elif isinstance(value, dict):
                        key_item = list(value.items())
                        value_key = key_item[0][0]  # 对应file2选择填充值的列名称
                        value_value = key_item[0][1]  # 对应file1选择填充的数量列表，列表的索引为file2的值顺序
                        # 按顺序填充对应数量的字段
                        if isinstance(value_value, list):
                            num = 0  # 修改file1的key对应值的次数，从0开始计数,即为file1对应key字段的下标
                            for index, index_value in enumerate(value_value):
                                file2_value = file2_data_frame[value_key].loc[
                                    index
                                ]  # 对应file2字段索引顺序的值
                                for k in range(index_value):
                                    file1_data_frame.loc[num, key] = file2_value  # 赋值
                                    num += 1
                        else:
                            raise Exception(f"{key_item}对应数据应为正整数列表")
                else:
                    raise Exception("缺少文件file2")
            else:
                raise Exception(f"{i}必须为字典数据")
        path = os.path.dirname(file1)  # file1同级文件目录
        file_name = "合并_" + os.path.basename(file1)
        save_file = os.path.join(path, file_name)
        file1_data_frame.to_excel(excel_writer=save_file, index=False)  # 存储到文件中
        return save_file

    def _sys_base_data_path(self):
        """根据运行用例的系统，返回对应环境的基础数据路径,data_type是基础数据类型"""
        host = os.environ["host"]
        data_path = os.path.join(ROOT_PATH, self.sys_name, "base_case_data", host.replace(":", "_"))
        if not os.path.exists(data_path):
            os.mkdir(data_path)
        return data_path

    def copy_base_file(self, file: str, name: str):
        """复制文件到基本数据目录"""
        data_path = self._sys_base_data_path()  # 构造基础数据的存储地址
        if not os.path.isfile(file):  # 判断文件是否存在
            raise FileNotFoundError(f"{file}文件未找到")
        else:
            new_file = os.path.join(data_path, name)  # 重命名存储文件名称
            shutil.copyfile(file, new_file)

    def base_data_path(self, file_name: str):
        """返回基本数据文件路径,如果存在文件则返回文件，否则抛出异常"""
        file = os.path.join(self._sys_base_data_path(), file_name)
        if os.path.isfile(file):
            return file
        else:
            return False

    def read_base_data_excel(self, file_name, sheet_name="", ignore_num: int = 0):
        """读取基础数据目录文件，并返回数据"""
        return self.read_excel(os.path.join(self._sys_base_data_path(), file_name), sheet_name, ignore_num)

    def read_all_sheet_func_save(self, file, replace_to=None, replace_data=None, ignore_num: int = 0):
        """调试中
        读取xls,xlsx数据，并将所有sheet页中的随机函数执行，并返回生成一个新的文件,如果需要替换数据
        @param file: 文件路径
        @param replace_to: 要替换的数据
        @param replace_data: 替换后的数据
        @param ignore_num:
        @return: """
        path = os.path.dirname(file)  # file同级文件目录
        file_name = "随机_" + os.path.basename(file)
        save_file = os.path.join(path, file_name)
        all_sheet = pandas.read_excel(file, sheet_name=None)
        with pandas.ExcelWriter(save_file) as writer:
            for single_sheet in list(all_sheet):
                data = self._read_excel_fun_data(file=file, sheet_name=single_sheet, ignore_num=ignore_num)
                if replace_to:
                    data.replace(replace_to, replace_data, inplace=True)
                data.to_excel(excel_writer=writer, index=False, sheet_name=single_sheet, startrow=ignore_num)
        return save_file


DjwPd = ExcelToolsPandas()  # 大教务的读取基础文件数据实例类
PlatFormPd = ExcelToolsPandas("platform_sys")  # 平台系统读取基础文件数据实例
ExamPd = ExcelToolsPandas("exam_sys")  # 考试系统读取基础文件数据实例
